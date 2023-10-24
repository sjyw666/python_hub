#研发者：时间遗忘
#开发时间：2023/8/16 20:53

import pandas as pd

# 读取Excel 文件(附件4)
fj4sj= pd.read_excel('附件4删除重复数据.xlsx')#print(fj4sp_id.head())#可以查看数据的前几行：

# 读取某一列的数据并转换为列表        #print(fj4sj_sp_id)# 打印列表
fj4sj_sp_id= fj4sj['sku_id'].tolist()#代码中的 '列名' 替换为您实际的第一列列名。这将使用 tolist() 方法将数据转换为 Python 的列表，然后将其存储在名为 first_column_data 的变量中。





# 读取Excel 文件(附件1)
fj1sj= pd.read_excel('附件1打折.xlsx')#print(fj1sj.head())#可以查看数据的前几行：

fj1sj_sp_id= fj1sj['sku_id'].tolist()

#print(len(fj1sj_sp_id))


# 指定要读取的行数和列数
start_row_index = 0                 # 读取的起始行索引
end_row_index = len(fj1sj_sp_id)    # 读取的结束行索引，在这里是计算长度将所有的进行遍历


start_col_index = 0                # 读取的起始列索引
end_col_index = 5                  # 读取的结束列索引




fj1sp_xx = []  # 创建存储信息的空列表(附件1中商品信息)

# 遍历所有行的所有单元格，将每行的第一个字符串数据存入子列表，然后将子列表存入总列表
for row_index in range(start_row_index, end_row_index):                     #遍历所有行

    #创建子列表
    cell_list = []

    first_dyg_sj = fj1sj.iloc[row_index, start_col_index]                #通过 iloc 属性获取指定行和列位置的数据，并将其存储在名为 fj1sysj 的变量中

    if isinstance(first_dyg_sj, str):
        cell_list.append(first_dyg_sj)

    for col_index in range(start_col_index + 1, end_col_index + 1):             #遍历所有单元格添加到子序列中,由于第一列先进行处理，则起始列加1
        dyg_sj = fj1sj.iloc[row_index, col_index]
        cell_list.append(dyg_sj)

    fj1sp_xx.append(cell_list)
'''
首先获取每一行的第一个单元格数据 first_dyg_sj。然后，我们检查这个数据是否为字符串类型，如果是字符串类型，则将其添加到 row_data 子列表中。

接着，我们使用内层循环遍历剩余的列，并将每个单元格的数据 cell_data 添加到 row_data 子列表中。最后，将每个子列表存储到 fj1sp_xx 总列表中。

这样，您将得到一个包含每一行的第一个字符串数据的子列表的总列表 fj1sp_xx。

if isinstance(first_dyg_sj, str): 是一个条件语句，用于检查 first_dyg_sj 变量的数据类型是否为字符串类型。

通过 isinstance() 函数判断数据类型，并返回布尔值，如果 first_dyg_sj 是字符串类型，则执行 row_data.append(first_dyg_sj) 将其添加到 row_data 子列表中。

这样，我们只将每行的第一个字符串类型数据添加到子列表中。如果第一个单元格的数据类型不是字符串，则忽略它，只将后续单元格的数据添加到子列表中。
'''

#print(len(fj1sp_xx))
''''''
import openpyxl
from datetime import datetime

# 创建新的 Excel 文件
workbook = openpyxl.Workbook()
worksheet = workbook.active

row_index = 1  # 行索引

for i in range(0, len(fj1sj_sp_id)):
    scores2 = {
        fj1sj_sp_id[i]: fj1sp_xx[i]
    }

    for item in fj4sj_sp_id:
        for item2 in scores2:
            if item == item2:
                every_ys = scores2.get(item2)

                for index, cell_value in enumerate(every_ys):
                    # 处理日期值
                    if isinstance(cell_value, datetime):
                        cell_value = cell_value.strftime("%Y-%m-%d")

                    # 写入单元格
                    worksheet.cell(row=row_index, column=index + 1, value=cell_value)

                row_index += 1

                # 检查行索引是否超出范围
                if row_index > 1048576:
                    # 超出范围，退出循环
                    break

        # 在内层循环结束后，再次检查行索引是否超出范围
        if row_index > 1048576:
            # 超出范围，退出循环
            break

    # 在最外层循环结束后，再次检查行索引是否超出范围
    if row_index > 1048576:
        # 超出范围，退出循环
        break

# 保存 Excel 文件
workbook.save("附件1分类后数据.xlsx")
#输出字典对应的value(商品信息)，将输出值赋予一个变量
#遍历次数为附件1商品的次数
#遍历列表元素(附件4商品id)
#遍历字典的值(附件1商品id)
#如果相等
'''''''''
使用 openpyxl 创建了一个新的 Excel 文件，并获取了默认的活动工作表。然后，我们使用嵌套循环遍历列表和字典，将匹配的元素写入到 Excel 文件中。

每匹配一次，我们使用 worksheet.cell() 方法将元素值写入到指定的行和列中，行索引 row_index 每次匹配后递增。最后，我们通过调用 workbook.save() 方法保存 Excel 文件为 “附件1分类后数据.xlsx”。

使用嵌套循环遍历这个子列表，并将每个元素按顺序依次放入单元格中。每次写入一个元素后，我们将行索引 row_index 递增，以便在下一行写入。

当子列表的所有元素都放入单元格后，外层循环会继续进行下一次迭代，从而写入下一个商品的信息。

子列表 every_ys  中的元素按顺序写入单元格。在这个循环中，cell_list_ys 变量代表子列表 every_ys 中的每个元素。

然后，我们使用 worksheet.cell() 方法将 item2 和 cell_list_ys  的值写入到指定的单元格中。

row_index 变量表示当前要写入的行索引，因此每写入一个元素，我们将行索引递增，以便将下一个元素写入到下一行。
'''

'''
workbook.save("output.xlsx") 将保存 Excel 文件到当前运行代码的工作目录中，也就是运行代码的当前目录。

当您运行这段代码时，生成的 Excel 文件 “output.xlsx” 将保存在与代码文件相同的目录下。如果您在运行代码的时候没有指定工作目录，则默认保存在运行代码的当前目录中。

您可以在文件资源管理器中检查当前目录来找到生成的 Excel 文件。如果您希望将文件保存到特定的目录中，可以在 "output.xlsx" 的路径部分提供完整的路径。

例如，如果您想将文件保存在特定目录 C:\Excel\output.xlsx 中，可以修改代码为 workbook.save("C:/Excel/output.xlsx")。

请确保代码执行时对写入文件的目录有写入权限。
'''